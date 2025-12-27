from flask import Flask, render_template, request, send_file, jsonify
import sqlite3
import smtplib
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import requests


app = Flask(__name__)

# ---------- DATABASE ----------
def get_db():
    return sqlite3.connect("attendance.db")

def init_db():
    db = get_db()
    cur = db.cursor()

    cur.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, name TEXT, email TEXT)")
    cur.execute("""CREATE TABLE IF NOT EXISTS attendance
                   (id INTEGER PRIMARY KEY, user_id INTEGER, date TEXT, status TEXT)""")

    # default user
    cur.execute("INSERT OR IGNORE INTO users VALUES (1,'Omkar','omkartalekar06@outlook.in')")
    db.commit()
    db.close()

init_db()

# ---------- ADD USER ----------
@app.route("/add_user", methods=["GET", "POST"])
def add_user():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]

        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO users (name, email) VALUES (?, ?)", (name, email))
        db.commit()
        db.close()

        return "User added successfully!"

    return render_template("add_user.html")


# ---------- MARK ATTENDANCE ----------
@app.route("/", methods=["GET", "POST"])
def mark():
    if request.method == "POST":
        user_id = request.form["user_id"]
        date = request.form["date"]
        status = request.form["status"]

        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO attendance (user_id, date, status) VALUES (?,?,?)",
                    (user_id, date, status))

        cur.execute("SELECT email FROM users WHERE id=?", (user_id,))
        email = cur.fetchone()[0]

        db.commit()
        db.close()

        return jsonify({"email": email, "date": date, "status": status})


    return render_template("mark.html")

# ---------- CALENDAR ----------
@app.route("/calendar/<int:user_id>")
def calendar(user_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("""SELECT attendance.date, users.name, attendance.status FROM attendance JOIN users ON attendance.user_id = users.id ORDER BY attendance.date""")
    data = cur.fetchall()
    db.close()

    return render_template("calendar.html", data=data)

# ---------- EXCEL REPORT ----------
@app.route("/report/<int:user_id>")
def report(user_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.append(["Date", "Status"])

    green = PatternFill(start_color="00FF00", fill_type="solid")
    red = PatternFill(start_color="FF0000", fill_type="solid")

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT date, status FROM attendance WHERE user_id=?", (user_id,))
    rows = cur.fetchall()
    db.close()

    for r in rows:
        ws.append(r)
        cell = ws.cell(row=ws.max_row, column=2)
        if r[1] == "Present":
            cell.fill = green
        else:
            cell.fill = red

    if not os.path.exists("reports"):
        os.mkdir("reports")

    path = f"reports/attendance_{user_id}.xlsx"
    wb.save(path)

    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
